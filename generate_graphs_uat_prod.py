from copy import copy
import csv
import os
import pandas as pd
import matplotlib.pyplot as plt
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from matplotlib.backends.backend_pdf import PdfPages

# Path to the folder where the CSV is saved
folder_name = 'reports'
csv_file_name = 'merged_prs_2024-09-24_to_2024-09-24'
csv_file_path = os.path.join('github', csv_file_name+'.csv')
excel_file_path = os.path.join(folder_name, 'pr_report.xlsx')
pdf_file_path = os.path.join(folder_name, 'pr_report_'+csv_file_name+'.pdf')
prod_release_count = 0
uat_release_count = 0

# Check if the CSV file exists
if not os.path.exists(csv_file_path):
    print(f"CSV file '{csv_file_path}' not found.")
    exit()

# Read the CSV and collect team names, release types, branches, and release titles
team_release_data = {}

with open(csv_file_path, mode='r', newline='', encoding='utf-8') as csv_file:
    reader = csv.DictReader(csv_file)

    for row in reader:
        team_name = row['Team Name']
        release_type = row['Release Type']
        branch = row['Branch']
        release_title = row['Title']

        # Normalize team and branch names as before
        if team_name.lower() == 'mpnl':
            team_name = 'coreportal'

        if team_name.lower() == 'fox':
            team_name = 'platypus'

        # Check if team is 'salescpq' and adjust based on title contents
        if team_name.lower() == 'salescpq':
            if 'sfka' in release_title.lower():
                team_name = 'koalas'
            elif 'sfjr' in release_title.lower():
                team_name = 'jaguar'

        # Only include valid team names
        if len(team_name) >= 3 and team_name != 'Unknown':
            if team_name not in team_release_data:
                team_release_data[team_name] = {
                    'total': 0,
                    'fast': 0,
                    'slow': 0,
                    'branches': set(),
                    'titles': [],
                    'prod': 0  # Initialize prod count
                }

            # Count prod and uat releases
            if branch == 'master':
                branch = 'prod'
                prod_release_count += 1
            elif branch == 'develop_uat':
                branch = 'uat'
                uat_release_count += 1

            # Track prod releases per team
            if branch == 'prod':
                team_release_data[team_name]['prod'] += 1  # Increment prod releases for the team

            # Increment counts for total, fast, and slow releases
            team_release_data[team_name]['total'] += 1
            if release_type == 'fast':
                team_release_data[team_name]['fast'] += 1
            else:
                team_release_data[team_name]['slow'] += 1

            # Keep track of the branches
            team_release_data[team_name]['branches'].add(branch)

            # Add release title to the list
            team_release_data[team_name]['titles'].append(release_title)




# Prepare data for Excel and PDF
excel_data = []
for team, data in team_release_data.items():
    # Join release titles with '\n' for line breaks
    releases_with_line_breaks = '\n'.join(data['titles'])
    excel_data.append([
        team,
        data['total'],
        data['fast'],
        data['slow'],
        ', '.join(data['branches']),  # Joined branch names with 'prod' or 'uat'
        releases_with_line_breaks
    ])

# Define column headers for Excel
columns = ['Team Name', 'Total Releases', 'Fast Releases', 'Slow Releases', 'Branches', 'Releases']

# Create a pandas DataFrame
df = pd.DataFrame(excel_data, columns=columns)

# Write the DataFrame to an Excel file
df.to_excel(excel_file_path, index=False, engine='openpyxl')

# Load the workbook to adjust cell formatting for line breaks
wb = load_workbook(excel_file_path)
ws = wb.active

# Ensure line breaks are enabled in the 'Releases' column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
    cell = row[0]
    cell.alignment = Alignment(wrapText=True)

# Save the updated Excel file with formatting
wb.save(excel_file_path)

# Continue from where your code generates the PDF with PdfPages(pdf_file_path)
with PdfPages(pdf_file_path) as pdf:
    # First page: Existing Bar Chart and Pie Chart
    # (Your original code remains unchanged for this part)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    # Bar chart: Number of releases by team
    teams = [team for team in team_release_data.keys()]
    pr_counts = [team_release_data[team]['total'] for team in teams]
    bars = ax1.bar(teams, pr_counts, color='orange')

    ax1.set_xticks(range(len(teams)))
    ax1.set_xticklabels(teams, rotation=45, ha='right')
    ax1.set_xlabel('Team')
    ax1.set_ylabel('Number of Releases')
    ax1.set_title('Number of Releases by Team')

    # Add release numbers inside each bar with larger font size
    for bar, count in zip(bars, pr_counts):
        height = bar.get_height()
        ax1.text(
            bar.get_x() + bar.get_width() / 2,  
            height - 1,  
            str(count),  
            ha='center',
            va='bottom', 
            fontsize=14,  
            color='white'
        )

    total_fast_releases = sum([entry['fast'] for entry in team_release_data.values()])
    total_slow_releases = sum([entry['slow'] for entry in team_release_data.values()])
    release_counts = [total_fast_releases, total_slow_releases]

    def absolute_numbers(val):
        total = sum(release_counts)
        return f'{int(round(val/100*total))}'  

    if total_fast_releases > 0 or total_slow_releases > 0:
        ax2.pie(release_counts, 
                labels=['Fast Releases', 'Slow Releases'], 
                autopct=absolute_numbers,
                startangle=90, 
                colors=['green', 'brown'])
        ax2.set_title('Fast vs Slow Releases')
    else:
        ax2.text(0.5, 0.5, 'No release data', ha='center', va='center', fontsize=12)

    plt.tight_layout()
    pdf.savefig()
    plt.close()

    # Second page: Table without "Releases" column
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.axis('off')

    table_data = [[team, data['total'], data['fast'], data['slow'], ', '.join(data['branches'])] for team, data in team_release_data.items()]
    col_labels = ['Team Name', 'Total', 'Fast', 'Slow', 'Branches']

    table = ax.table(cellText=table_data,
                     colLabels=col_labels,
                     loc='center',
                     cellLoc='center',
                     colWidths=[0.2, 0.1, 0.1, 0.1, 0.2])
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 1.2)
    plt.tight_layout()
    pdf.savefig()
    plt.close()

    # Create a page for each team with their data and release titles
    for team, data in team_release_data.items():
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.axis('off')

        ax.text(0, 0.9, f'Team: {team}', fontsize=14, fontweight='bold', ha='left')
        ax.text(0, 0.8, f'Total Releases: {data["total"]}', fontsize=12, ha='left')
        ax.text(0, 0.7, f'Fast Releases: {data["fast"]}', fontsize=12, ha='left')
        ax.text(0, 0.6, f'Slow Releases: {data["slow"]}', fontsize=12, ha='left')
        ax.text(0, 0.5, f'Branches: {", ".join(data["branches"])}', fontsize=12, ha='left')

        ax.text(0, 0.4, 'Releases:', fontsize=12, fontweight='bold', ha='left')
        for idx, title in enumerate(data['titles'], start=1):
            ax.text(0.05, 0.4 - idx * 0.05, f'{idx}. {title}', fontsize=10, ha='left')

        plt.tight_layout()
        pdf.savefig()
        plt.close()

    # Additional Pages with New Graphs

    # Page: Fast vs. Slow Releases by Team (Side-by-side Bar Chart)
    # Page: Number of Prod vs UAT Releases by Team (Side-by-side Bar Chart)
    fig, ax = plt.subplots(figsize=(10, 6))

    prod_counts = [team_release_data[team]['prod'] for team in teams]
    uat_counts = [team_release_data[team]['total'] - team_release_data[team]['prod'] for team in teams]  # Total minus prod gives UAT count

    bar_width = 0.35
    index = range(len(teams))

    # Bars for prod releases (colored purple)
    bars_prod = ax.bar(index, prod_counts, bar_width, label='Prod Releases', color='purple')

    # Bars for uat releases (colored orange), shifted by bar_width for side-by-side appearance
    bars_uat = ax.bar([i + bar_width for i in index], uat_counts, bar_width, label='UAT Releases', color='orange')

    # Set x-axis labels and title
    ax.set_xticks([i + bar_width / 2 for i in index])
    ax.set_xticklabels(teams, rotation=45, ha='right')
    ax.set_xlabel('Team')
    ax.set_ylabel('Number of Releases')
    ax.set_title('Number of Prod vs UAT Releases by Team')

    # Add legend for the bars
    ax.legend()

    # Add release numbers inside each bar
    for bar, count in zip(bars_prod, prod_counts):
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,  
            height - 1,  
            str(count),  
            ha='center',
            va='bottom', 
            fontsize=12,  
            color='white'
        )

    for bar, count in zip(bars_uat, uat_counts):
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,  
            height - 1,  
            str(count),  
            ha='center',
            va='bottom', 
            fontsize=12,  
            color='white'
        )

    # Calculate the total number of prod releases across all teams
    total_prod_releases = sum(prod_counts)

    # Add a text annotation to display the total number of prod releases for the week
    ax.text(0.95, 0.95, f'Total Prod Releases this week: {total_prod_releases}', 
            transform=ax.transAxes, fontsize=14, fontweight='bold', ha='right', va='top', color='black')

    plt.tight_layout()
    pdf.savefig()
    plt.close()


print(f"PDF file saved to {pdf_file_path}")

print(f"PDF file saved to {pdf_file_path}")
print(f"Excel file saved to {excel_file_path}")
